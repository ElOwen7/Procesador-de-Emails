import os
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime, timedelta
import zipfile
from io import BytesIO
from correo_utils import conectar_gmail, obtener_correos

CARPETA_ADJUNTOS = "Adjuntos"
if not os.path.exists(CARPETA_ADJUNTOS):
    os.makedirs(CARPETA_ADJUNTOS)

st.set_page_config(page_title="Procesador de Correos", layout="wide")
st.title("📥 Procesador de Correos de Vacantes")

# --- Login UI ---
st.subheader("🔐 Iniciar Sesión")
with st.form("login_form"):
    user_mail = st.text_input("Correo electrónico", placeholder="tu_email@gmail.com")
    password_app = st.text_input("Contraseña de aplicación", type="password", placeholder="************")
    login_button = st.form_submit_button("Conectar")

if login_button:
    try:
        imap = conectar_gmail(user_mail, password_app)
        imap.logout()
        st.session_state["logged_in"] = True
        st.session_state["user_mail"] = user_mail
        st.session_state["password_app"] = password_app
        st.success("✅ Conexión exitosa. Puedes continuar.")
    except Exception as e:
        st.session_state["logged_in"] = False
        st.error("❌ Error al conectar. Revisa tu correo o contraseña.")

# --- Si conexión fue exitosa, mostrar filtros ---
if st.session_state.get("logged_in"):
    st.subheader("📨 Filtrado de Correos")

    filtro_asunto = st.text_input("🔍 Filtro de asunto", value="VACANTE")
    fecha_desde = st.date_input("📅 Fecha desde", value=datetime.now() - timedelta(days=7))

    if st.button("🔄 Procesar correos"):
        st.info("Conectando con Gmail y procesando correos...")

        imap = conectar_gmail(st.session_state["user_mail"], st.session_state["password_app"])
        datos = obtener_correos(imap, filtro_asunto, fecha_desde, CARPETA_ADJUNTOS)
        imap.logout()

        if datos:
            df = pd.DataFrame(datos)
            st.session_state["df_correo"] = df
            st.success("✅ Correos procesados correctamente.")
        else:
            st.warning("⚠️ No se encontraron correos que coincidan con el filtro.")

# --- Mostrar resultados ---
if "df_correo" in st.session_state:
    df = st.session_state["df_correo"]
    st.subheader("📋 Resultados")
    st.dataframe(df.drop(columns=["attachment_path"]), use_container_width=True)

    # CSV
    df_csv = df.drop(columns=["attachment_path"])
    csv = df_csv.to_csv(index=False).encode("utf-8")
    st.download_button("⬇️ Descargar CSV", data=csv, file_name="correos_filtrados.csv", mime="text/csv")

    # Archivos individuales
    st.subheader("📎 Archivos Adjuntos")
    for i, row in df.iterrows():
        if row["attachment_path"] and os.path.exists(row["attachment_path"]):
            with open(row["attachment_path"], "rb") as f:
                bytes_data = f.read()
            st.download_button(
                label=f"📄 Descargar {row['attachment_name']}",
                data=BytesIO(bytes_data),
                file_name=row["attachment_name"],
                mime="application/octet-stream",
                key=f"adjunto_{i}"
            )

    # ZIP
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # Crear Excel con hipervínculos
        wb = Workbook()
        ws = wb.active
        ws.title = "Correos"

        df_excel = df.copy()

        # Crear encabezado
        for col_idx, col_name in enumerate(df_excel.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

        # Llenar datos
        for row_idx, row in enumerate(df_excel.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                col_name = df_excel.columns[col_idx - 1]

                if col_name == "attachment_name" and getattr(row, "attachment_path"):
                    link = f"adjuntos/{getattr(row, 'attachment_name')}"
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = getattr(row, "attachment_name")
                    cell.hyperlink = link
                    cell.style = "Hyperlink"
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Guardar Excel en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Agregar Excel al ZIP
        zip_file.writestr("correos_filtrados.xlsx", excel_buffer.read())

        # Agregar adjuntos
        for i, row in df.iterrows():
            if row["attachment_path"] and os.path.exists(row["attachment_path"]):
                with open(row["attachment_path"], "rb") as f:
                    file_data = f.read()
                    arcname = f"adjuntos/{row['attachment_name']}"
                    zip_file.writestr(arcname, file_data)

    zip_buffer.seek(0)

    st.download_button(
        label="📦 Descargar ZIP (Excel + Adjuntos)",
        data=zip_buffer,
        file_name="correos_y_adjuntos.zip",
        mime="application/zip"
    )
